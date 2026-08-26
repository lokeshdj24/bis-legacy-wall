from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys


def get_documents_dir() -> Path:
    if sys.platform == "win32":
        import ctypes
        from ctypes import wintypes

        buf = ctypes.create_unicode_buffer(wintypes.MAX_PATH)
        ctypes.windll.shell32.SHGetFolderPathW(None, 5, None, 0, buf)
        return Path(buf.value)
    return Path.home() / "Documents"


data_dir = get_documents_dir() / "legacy_wall"
images_dir = data_dir / "images"
data_dir.mkdir(parents=True, exist_ok=True)
images_dir.mkdir(parents=True, exist_ok=True)

wb = Workbook()
header_fill = PatternFill("solid", fgColor="0B1B3A")
header_font = Font(bold=True, color="FFCA1B")

# ---------- Cards (Sheet 1) — eras derived from source year ranges ----------
cards = wb.active
cards.title = "Cards"
card_headers = ["card_id", "sort_order", "title", "era", "image"]
cards.append(card_headers)
for col, _ in enumerate(card_headers, 1):
    cell = cards.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Dummy titles/images for now; eras taken from the provided source pages
card_rows = [
    [1, 1, "Card Title 1", "1946 - 1948", "card-01-cover.jpg"],
    [2, 2, "Card Title 2", "1949 - 1960", "card-02-cover.jpg"],
    [3, 3, "Card Title 3", "1961 - 1969", "card-03-cover.jpg"],
    [4, 4, "Card Title 4", "1972 - 1997", "card-04-cover.jpg"],
]
for row in card_rows:
    cards.append(row)

# ---------- Slides (Sheet 2) ----------
slides = wb.create_sheet("Slides")
slide_headers = [
    "slide_id",
    "card_id",
    "year",
    "sort_order",
    "eyebrow",
    "title",
    "description",
    "image",
]
slides.append(slide_headers)
for col, _ in enumerate(slide_headers, 1):
    cell = slides.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# description = short UI caption (italic in source); falls back to archive text when caption missing
slide_rows = [
    # ---- Card 1: 1946 - 1948 ----
    [
        1,
        1,
        1946,
        1,
        "3 SEPTEMBER 1946",
        "India’s Standards Body Is Born",
        "The Indian Standards Institution is established.",
        "1946-01.jpg",
    ],
    [
        2,
        1,
        1947,
        1,
        "JANUARY 1947",
        "The ISI Constitution Adopted",
        "India's standards framework set before its freedom.",
        "1947-01.jpg",
    ],
    [
        3,
        1,
        1947,
        2,
        "6 JANUARY 1947",
        "The First General Council Convenes",
        "John Mathai chairs as First President of ISI.",
        "1947-02.jpg",
    ],
    [
        4,
        1,
        1947,
        3,
        "15 AUGUST 1947",
        "India Becomes Free. ISI Is Already at Work.",
        "Standards were not an afterthought of nation-building.",
        "1947-03.jpg",
    ],
    [
        5,
        1,
        1948,
        1,
        "FEBRUARY 1948 · CALCUTTA",
        "The Birth of India’s Quality Movement",
        "Shewhart, Mahalanobis and Verman in one founding photograph.",
        "1948-01.jpg",
    ],
    [
        6,
        1,
        1948,
        2,
        "1948 · CALCUTTA",
        "A Governor-General Speaks for Standards",
        "Rajagopalachari bridges political leadership and standardization.",
        "1948-02.jpg",
    ],
    [
        7,
        1,
        1948,
        3,
        "29 SEPTEMBER 1948",
        "Government Commits to Standards",
        "Dr Mookerjee presides — political support for industrialization.",
        "1948-03.jpg",
    ],
    # ---- Card 2: 1949 - 1960 ----
    [
        8,
        2,
        1949,
        1,
        "1949",
        "A Plot of Land. A Future Home.",
        "The site where Manak Bhawan will be built.",
        "1949-01.jpg",
    ],
    [
        9,
        2,
        1949,
        2,
        "JUNE–JULY 1949 · PARIS",
        "India at the ISO Plenary",
        "Dr Verman represents ISI — two years after Independence.",
        "1949-02.jpg",
    ],
    [
        10,
        2,
        1950,
        1,
        "JULY 1950",
        "Vice-President of ISO",
        "Dr Verman leads India at the 4th ISO Council Meeting.",
        "1950-01.jpg",
    ],
    [
        11,
        2,
        1951,
        1,
        "MARCH 1951",
        "PM Nehru Visits ISI",
        "Indira Gandhi present — two generations of leadership in one frame.",
        "1951-01.jpg",
    ],
    [
        12,
        2,
        1951,
        2,
        "10 AUGUST 1951",
        "IS:1-1951 — The National Flag Is a Standard",
        "Nehru receives the first Tricolour made to IS:1.",
        "1951-02.jpg",
    ],
    [
        13,
        2,
        1952,
        1,
        "1952",
        "ISI’s Humble Beginning",
        "The top floor of Shri Ram Institute — ISI’s first office.",
        "1952-01.jpg",
    ],
    [
        14,
        2,
        1954,
        1,
        "21 AUGUST 1954 · MANAK BHAWAN",
        "Nehru Lays the Foundation",
        "India's standards body gets its permanent home.",
        "1954-01.jpg",
    ],
    [
        15,
        2,
        1955,
        1,
        "16 JULY 1955",
        "The First ISI Mark Licence",
        "Every ISI and BIS mark in India traces to this moment.",
        "1955-01.jpg",
    ],
    [
        16,
        2,
        1958,
        1,
        "1958 · MANAK BHAWAN",
        "India’s Standards Home Inaugurated",
        "Nehru, Shastri, and Satish Gujral — art and standards in one frame.",
        "1958-01.jpg",
    ],
    [
        17,
        2,
        1960,
        1,
        "1960 · NEW DELHI",
        "First IEC Meeting in Asia",
        "Dr Radhakrishnan inaugurates — India leads global electrical standards.",
        "1960-01.jpg",
    ],
    # ---- Card 3: 1961 - 1969 ----
    [
        18,
        3,
        1961,
        1,
        "1961",
        "ISI Tells Its Story",
        "Display boards show the institution’s growing reach and standards.",
        "1961-01.jpg",
    ],
    [
        19,
        3,
        1962,
        1,
        "1962",
        "Every Flag, Certified",
        "National flags stamped with the ISI mark — IS:1-1951 in action.",
        "1962-01.jpg",
    ],
    [
        20,
        3,
        1962,
        2,
        "1962",
        "The President Examines the Mark",
        "Dr Rajendra Prasad inspects ISI’s certification scheme.",
        "1962-02.jpg",
    ],
    [
        21,
        3,
        1964,
        1,
        "3 AUGUST 1964",
        "The Vice-President at Manak Bhawan",
        "Dr Zakir Hussain visits — standards eliminate costly experiments.",
        "1964-01.jpg",
    ],
    [
        22,
        3,
        1966,
        1,
        "FEBRUARY 1966",
        "Two Eminent Scientists at ISI",
        "Dr Narlikar and Dr Verman — the scientific ecosystem of ISI.",
        "1966-01.jpg",
    ],
    [
        23,
        3,
        1967,
        1,
        "7 JUNE 1967",
        "New Leadership, Same Mission",
        "Dr A.N. Ghosh succeeds Dr Verman as Director-General of ISI.",
        "1967-01.jpg",
    ],
    [
        24,
        3,
        1969,
        1,
        "1969",
        "Two Future Presidents at ISI",
        "A remarkable photograph: two individuals who would each serve as President of India, walking together at an ISI Governing Council meeting.",
        "1969-01.jpg",
    ],
    # ---- Card 4: 1972 - 1997 ----
    [
        25,
        4,
        1972,
        1,
        "DECEMBER 1972 · CALCUTTA",
        "ISI Silver Jubilee",
        "Twenty-five years of standards — celebrated in the city of quality.",
        "1972-01.jpg",
    ],
    [
        26,
        4,
        1972,
        2,
        "~1972",
        "A Future President at ISRO",
        "The young Abdul Kalam welcomes ISI visitors to ISRO.",
        "1972-02.jpg",
    ],
    [
        27,
        4,
        1976,
        1,
        "13 JANUARY 1976",
        "Agriculture Enters the Standards Mandate",
        "Dr Swaminathan links standards to India’s food security revolution.",
        "1976-01.jpg",
    ],
    [
        28,
        4,
        1976,
        2,
        "12 JULY 1976 · MADRAS",
        "ISI Comes to Southern India",
        "A new regional office and laboratory complex inaugurated.",
        "1976-02.jpg",
    ],
    [
        29,
        4,
        1986,
        1,
        "23 DECEMBER 1986",
        "Parliament Creates the Bureau of Indian Standards",
        "The BIS Act 1986 — ISI’s mission reborn in law.",
        "1986-01.jpg",
    ],
    [
        30,
        4,
        1987,
        1,
        "18 APRIL 1987",
        "The Bureau of Indian Standards Convenes",
        "Born before Independence. Reborn by Parliament. Same mission, new mandate.",
        "1987-01.jpg",
    ],
    [
        31,
        4,
        1993,
        1,
        "20 AUGUST 1993",
        "India’s Highest Quality Award Presented",
        "PM Narasimha Rao presents the first Rajiv Gandhi National Quality Award.",
        "1993-01.jpg",
    ],
    [
        32,
        4,
        1997,
        1,
        "15 SEPTEMBER 1997",
        "BIS Golden Jubilee",
        "Fifty years of standards — PM Gujral inaugurates at Vigyan Bhawan.",
        "1997-01.jpg",
    ],
]

for row in slide_rows:
    slides.append(row)

# ---------- README ----------
guide = wb.create_sheet("README")
guide["A1"] = "BIS Legacy Wall — Content Guide"
guide["A1"].font = Font(bold=True, size=14, color="0B1B3A")
guide.merge_cells("A1:B1")
guide["A2"] = "Topic"
guide["B2"] = "Details"
guide["A2"].font = Font(bold=True)
guide["B2"].font = Font(bold=True)

guide_lines = [
    ("", ""),
    ("Location", r"Documents\legacy_wall\ (dynamic per Windows user via app.getPath('documents'))"),
    ("Excel", "legacy-wall-content.xlsx"),
    ("Images", r"Documents\legacy_wall\images\ — filenames only in Excel columns"),
    ("", ""),
    ("Cards sheet", "One row = one card on the Card List (title, era, cover image)"),
    ("Slides sheet", "One row = one gallery story/image; many rows can share card_id + year"),
    ("card_id", "Links slides to a card. Opening a card shows only matching slides."),
    ("year", "Builds gallery timeline buttons for that card"),
    ("sort_order", "Order of images within the same year"),
    ("eyebrow / title / description", "Gallery text fields"),
]
for i, (a, b) in enumerate(guide_lines, start=3):
    guide.cell(i, 1, a)
    guide.cell(i, 2, b)

for i, w in enumerate([10, 12, 28, 14, 28], 1):
    cards.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate([10, 10, 10, 12, 28, 48, 70, 22], 1):
    slides.column_dimensions[get_column_letter(i)].width = w
guide.column_dimensions["A"].width = 28
guide.column_dimensions["B"].width = 90

out = data_dir / "legacy-wall-content.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Cards: {len(card_rows)}  Slides: {len(slide_rows)}")
print(f"Images folder: {images_dir}")
